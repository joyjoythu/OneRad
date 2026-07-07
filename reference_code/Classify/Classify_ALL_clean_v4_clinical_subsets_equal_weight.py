"""
Medical Image Classification with Radiomics and Deep Learning Features.

A refactored classifier combining radiomics features and CLIP-based deep learning 
features for medical image classification tasks.

Upgraded version: additionally outputs performance statistics by clinical indicator subsets (ER, PR, Her2, Ki67).

Author: Atsea
Created: 2021/12/22
Refactored: 2025
Upgraded: 2026/06/15
Modified: 2026/06/16 - Task 0 and Task 6 use equal fusion weights (0.5 each)
"""

import datetime
import gc
import os
import random
import shutil
import time
import warnings
import shap
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import h5py
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np
import pandas as pd
import torch
from matplotlib import rcParams
from openpyxl import Workbook, load_workbook
from sklearn import feature_selection, metrics, preprocessing
from sklearn.ensemble import GradientBoostingClassifier, RandomForestClassifier
from sklearn.linear_model import Lasso, LassoCV, LogisticRegression
from sklearn.model_selection import GridSearchCV, StratifiedKFold
from sklearn.svm import SVC

warnings.filterwarnings('ignore')


# ==============================================================================
# Data Classes
# ==============================================================================

@dataclass
class ClassifierConfig:
    """Configuration for classifier parameters."""
    # Data paths
    save_path: Path
    radio_feature_path_0: Path
    clip_feature_path_0: Path
    radio_feature_path_6: Path
    clip_feature_path_6: Path
    ser_path: Path
    feature_excel_path: Optional[Path] = None
    
    # Task settings
    tasks: List[str] = field(default_factory=lambda: ['0', '6'])
    data_types: List[str] = field(default_factory=lambda: ['clip', 'radio'])
    
    # Validation settings
    use_external_test: bool = True
    k_fold: int = 5
    cv_random_state: int = 112
    grid_random_state: int = 7456
    
    # Feature selection settings
    distill_mode: str = 'mean'
    use_lassocv: bool = True
    lassocv_k: int = 30
    select_kbest_k: int = 12
    select_kbest_func: int = 1
    
    # Classifier settings
    use_lr: bool = True
    use_svm: bool = False
    use_rf: bool = False
    use_gdbt: bool = False
    
    # Fusion settings
    fusion_min_improvement: float = 0.0  # min AUC improvement over Task6-only to use fusion


@dataclass  
class MetricsResult:
    """Container for classification metrics."""
    accuracy: float = 0.0
    sensitivity: float = 0.0
    specificity: float = 0.0
    auc: float = 0.0
    best_threshold: float = 0.5
    tp: int = 0
    tn: int = 0
    fn: int = 0
    fp: int = 0


@dataclass
class PatientData:
    """Container for patient feature data."""
    patient_id: str
    clip_feature: np.ndarray
    radio_feature: np.ndarray
    label: int
    clinical_features: Optional[np.ndarray] = None
    ser_feature: Optional[np.ndarray] = None


# ==============================================================================
# Utility Functions
# ==============================================================================

def set_plot_config():
    """Set matplotlib configuration for Chinese and math text."""
    config = {
        "font.family": 'Times New Roman',
        "font.size": 13,
        "mathtext.fontset": 'stix',
        "font.serif": ['SimSun'],
    }
    rcParams.update(config)
    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['axes.unicode_minus'] = False


def ensure_dir(path: Path):
    """Create directory if not exists, remove and recreate if exists."""
    if path.exists():
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)


def shuffle_array(x: np.ndarray, axis: int = 0) -> np.ndarray:
    """Shuffle array along specified axis."""
    new_index = list(range(x.shape[axis]))
    random.shuffle(new_index)
    
    dims = list(range(len(x.shape)))
    dims[0], dims[axis] = dims[axis], dims[0]
    x_new = np.transpose(x, dims)
    x_new = x_new[new_index]
    return np.transpose(x_new, dims)


def normalize_features(
    train_data: np.ndarray, 
    test_data: Optional[np.ndarray] = None,
    method: str = 'StandardScaler'
) -> Tuple[np.ndarray, Optional[np.ndarray]]:
    """Normalize features using specified method."""
    if method == 'MinMaxScaler':
        scaler = preprocessing.MinMaxScaler()
    elif method == 'StandardScaler':
        scaler = preprocessing.StandardScaler()
    else:
        raise ValueError(f"Unknown method: {method}")
    
    train_normalized = scaler.fit_transform(train_data)
    test_normalized = scaler.transform(test_data) if test_data is not None else None
    return train_normalized, test_normalized


def calculate_metrics(
    y_true: np.ndarray, 
    y_prob: np.ndarray,
    threshold: Optional[float] = None
) -> MetricsResult:
    """Calculate classification metrics."""
    result = MetricsResult()
    
    if len(np.unique(y_true)) < 2:
        result.auc = 0.0
    else:
        result.auc = metrics.roc_auc_score(y_true, y_prob)
    
    fpr, tpr, thresholds = metrics.roc_curve(y_true, y_prob)
    youden_index = tpr + (1 - fpr)
    result.best_threshold = thresholds[np.argmax(youden_index)]
    
    if result.best_threshold > 1:
        result.best_threshold = 0.5
    if threshold is not None:
        result.best_threshold = threshold
    
    y_pred = (y_prob >= result.best_threshold).astype(int)
    
    result.tp = int(np.sum((y_true == 1) & (y_pred == 1)))
    result.tn = int(np.sum((y_true == 0) & (y_pred == 0)))
    result.fp = int(np.sum((y_true == 0) & (y_pred == 1)))
    result.fn = int(np.sum((y_true == 1) & (y_pred == 0)))
    
    result.sensitivity = result.tp / (result.tp + result.fn + 1e-16)
    result.specificity = result.tn / (result.tn + result.fp + 1e-16)
    result.accuracy = (result.tp + result.tn) / (result.tp + result.tn + result.fp + result.fn + 1e-16)
    
    return result


def tfn_fusion(first_feature: np.ndarray, second_feature: np.ndarray) -> np.ndarray:
    """Tensor Fusion Network feature fusion."""
    first = torch.from_numpy(first_feature).float()
    second = torch.from_numpy(second_feature).float()
    n = first.shape[0]
    
    A = torch.cat([first, torch.ones(n, 1)], dim=1)
    B = torch.cat([second, torch.ones(n, 1)], dim=1)
    
    A = A.unsqueeze(2)
    B = B.unsqueeze(1)
    fusion = torch.einsum('nxt,nty->nxy', A, B)
    
    return fusion.flatten(start_dim=1).numpy()


# ==============================================================================
# Excel Utilities
# ==============================================================================

def write_to_excel(
    path: Path, 
    data: Any, 
    row: int, 
    col: int, 
    sheet_name: str = 'Sheet'
):
    """Write single cell to Excel file."""
    path_str = str(path)
    
    if path.exists():
        wb = load_workbook(path_str)
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(sheet_name)
        sheet.cell(row, col, data)
        wb.save(path_str)
        wb.close()
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.title = sheet_name
        sheet.cell(row, col, data)
        wb.save(path_str)
        wb.close()


def save_grid_params(
    path: Path, 
    iteration: int, 
    model, 
    classifier_name: str
):
    """Save grid search parameters to Excel."""
    if iteration == 1:
        write_to_excel(path, 'Round', 1, 1, classifier_name)
    
    write_to_excel(path, str(iteration), iteration + 1, 1, classifier_name)
    
    col = 1
    for param_name in model.best_params_.keys():
        col += 1
        if iteration == 1:
            write_to_excel(path, str(param_name), 1, col, classifier_name)
        write_to_excel(path, str(model.best_params_[param_name]), iteration + 1, col, classifier_name)


# ==============================================================================
# Feature Selection
# ==============================================================================

class FeatureSelector:
    """Feature selection using various methods."""
    
    SELECT_FUNCTIONS = {
        1: feature_selection.f_classif,
        2: feature_selection.mutual_info_classif,
        3: feature_selection.chi2,
        4: feature_selection.f_regression,
        5: feature_selection.mutual_info_regression,
    }
    
    def __init__(self, feature_names: Optional[np.ndarray] = None):
        self.feature_names = feature_names
        self.selected_indices = None
        self.coef_ = None
        self.intercept_ = None
    
    def select_k_best(
        self, 
        X: np.ndarray, 
        y: np.ndarray, 
        k: int, 
        func_id: int = 1
    ) -> Tuple[np.ndarray, np.ndarray]:
        """Select k best features using specified function."""
        func = self.SELECT_FUNCTIONS.get(func_id, feature_selection.mutual_info_classif)
        selector = feature_selection.SelectKBest(func, k=k)
        X_selected = selector.fit_transform(X, y)
        self.selected_indices = selector.get_support()
        return X_selected, self.selected_indices
    
    def lasso_select(
        self, 
        X: np.ndarray, 
        y: np.ndarray, 
        alpha: float = 0.0008
    ) -> np.ndarray:
        """Feature selection using LASSO."""
        model = Lasso(alpha=alpha, max_iter=500).fit(X, y)
        return model.coef_
    
    def lasso_cv_select(
        self, 
        X: np.ndarray, 
        y: np.ndarray,
        save_path: Path,
        n_alphas: int = 30,
        feature_names: Optional[List[str]] = None
    ) -> Tuple[np.ndarray, float]:
        """Feature selection using LASSO with cross-validation."""
        set_plot_config()
        
        alphas = np.logspace(-6, 1, n_alphas)
        model = LassoCV(alphas=alphas, cv=10, max_iter=100000, fit_intercept=True).fit(X, y)
        
        coef = model.coef_
        intercept = model.intercept_
        self.coef_ = coef
        self.intercept_ = intercept
        
        self._plot_lasso_path(model, X, y, save_path)
        self._plot_mse(model, save_path)
        
        # Save feature weights and plot
        if feature_names is not None:
            self._save_lasso_weights(coef, intercept, feature_names, save_path)
            plot_feature_weights(coef, save_path, feature_names)
        
        return coef, intercept
    
    def _plot_lasso_path(self, model, X, y, save_path: Path):
        """Plot LASSO regularization path."""
        alphas = np.logspace(-6, 1, 30)
        coefs = model.path(X, y, alphas=alphas, max_iter=100000)[1].T
        
        plt.figure()
        plt.semilogx(model.alphas_, coefs, '-')
        plt.axvline(model.alpha_, color='black', ls='--')
        plt.xlabel('Lambda')
        plt.ylabel('Coefficient')
        plt.savefig(save_path / 'lassocv.png', dpi=600)
        plt.close()
    
    def _plot_mse(self, model, save_path: Path):
        """Plot MSE curve."""
        mse_mean = np.mean(model.mse_path_, axis=1)
        mse_std = np.std(model.mse_path_, axis=1)
        
        plt.figure()
        plt.errorbar(model.alphas_, mse_mean, yerr=mse_std, fmt='o', 
                     ms=3, mfc='r', mec='r', ecolor='lightblue',
                     elinewidth=2, capsize=4, capthick=1)
        plt.semilogx()
        plt.axvline(model.alpha_, color='black', ls='--')
        plt.xlabel('Lambda')
        plt.ylabel('MSE')
        plt.savefig(save_path / 'mse.png', dpi=600)
        plt.close()
    
    def _save_lasso_weights(self, coef: np.ndarray, intercept: float,
                           feature_names: List[str], save_path: Path):
        """Save LASSO feature weights to Excel."""
        try:
            weight_path = save_path / 'lassocv_feature_weight.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = 'Sheet'
            ws.cell(1, 1, '特征名')
            ws.cell(1, 2, '特征权重')
            ws.cell(1, 3, '截止值')
            ws.cell(2, 3, float(intercept))
            
            for i, name in enumerate(feature_names):
                ws.cell(i + 2, 1, str(name))
                ws.cell(i + 2, 2, float(coef[i]))
            
            wb.save(weight_path)
            wb.close()
            print(f"LASSO weights saved to {weight_path}")
        except Exception as e:
            print(f"Error saving LASSO weights: {e}")


DEFAULT_SUBSET_INDICATORS = {
    'ER': 'ER（阳性为1）',
    'PR': 'PR（阳性为1）',
    'Her2': 'Her2（阳性为1）',
    'Ki67': 'Ki67（高表达为1）',
}

SUBSET_VALUE_LABELS = {
    'ER': {0: '阴性', 1: '阳性'},
    'PR': {0: '阴性', 1: '阳性'},
    'Her2': {0: '阴性', 1: '阳性'},
    'Ki67': {0: '低表达', 1: '高表达'},
}


def find_clinical_column(df: pd.DataFrame, keywords: List[str], exclude: Optional[str] = None) -> Optional[str]:
    """Find column name by keywords, optional exclusion."""
    matches = []
    for col in df.columns:
        col_lower = col.lower()
        if all(kw.lower() in col_lower for kw in keywords):
            if exclude and exclude.lower() in col_lower:
                continue
            matches.append(col)
    return matches[0] if matches else None


def aggregate_fold_predictions(
    ids: np.ndarray,
    probs: np.ndarray,
    labels: np.ndarray
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Aggregate repeated predictions per patient ID by averaging probabilities."""
    df = pd.DataFrame({'ID': ids, 'Prob': probs, 'Label': labels})
    agg = df.groupby('ID').agg({'Prob': 'mean', 'Label': 'first'}).reset_index()
    return agg['ID'].values, agg['Prob'].values, agg['Label'].values


class ClinicalSubsetAnalyzer:
    """Analyze classification metrics stratified by clinical indicators."""

    def __init__(
        self,
        clinical_excel_path: Path,
        id_col: str = 'ID',
        indicators: Optional[Dict[str, str]] = None
    ):
        self.clinical_df = pd.read_excel(clinical_excel_path)
        self.id_col = id_col
        self.indicators = indicators or DEFAULT_SUBSET_INDICATORS
        self._resolve_columns()

    def _resolve_columns(self):
        """Resolve actual column names for ID and indicators."""
        if self.id_col not in self.clinical_df.columns:
            found = find_clinical_column(self.clinical_df, ['id'])
            if found:
                self.id_col = found

        self.resolved_indicators = {}
        for name, expected_col in self.indicators.items():
            if expected_col in self.clinical_df.columns:
                self.resolved_indicators[name] = expected_col
            else:
                keywords = [name.lower()]
                exclude_kw = 'her' if name.lower() == 'er' else None
                found = find_clinical_column(self.clinical_df, keywords, exclude=exclude_kw)
                if found:
                    self.resolved_indicators[name] = found
                else:
                    print(f"Warning: clinical indicator column not found for {name}")

        self.clinical_df[self.id_col] = self.clinical_df[self.id_col].astype(str)

    def get_indicator_values(self, patient_ids: np.ndarray, indicator_name: str) -> np.ndarray:
        """Get indicator values for given patient IDs."""
        if indicator_name not in self.resolved_indicators:
            return np.array([np.nan] * len(patient_ids))

        col = self.resolved_indicators[indicator_name]
        values = []
        for pid in patient_ids:
            pid_str = str(pid)
            row = self.clinical_df[self.clinical_df[self.id_col] == pid_str]
            if row.empty:
                print(f"Warning: patient {pid} not found in clinical data")
                values.append(np.nan)
            else:
                values.append(row[col].values[0])

        return np.array(values, dtype=float)

    def analyze(
        self,
        patient_ids: np.ndarray,
        probs: np.ndarray,
        labels: np.ndarray,
        threshold: float = 0.5
    ) -> pd.DataFrame:
        """Compute metrics for each clinical indicator subset."""
        results = []

        for indicator_name in self.resolved_indicators.keys():
            indicator_values = self.get_indicator_values(patient_ids, indicator_name)
            valid_mask = ~np.isnan(indicator_values)

            if not np.any(valid_mask):
                print(f"Warning: no valid samples for indicator {indicator_name}")
                continue

            valid_ids = patient_ids[valid_mask]
            valid_probs = probs[valid_mask]
            valid_labels = labels[valid_mask]
            valid_values = indicator_values[valid_mask]

            unique_values = np.unique(valid_values)

            for val in unique_values:
                subset_mask = valid_values == val
                subset_probs = valid_probs[subset_mask]
                subset_labels = valid_labels[subset_mask]
                n_samples = len(subset_labels)

                if n_samples == 0:
                    continue

                if len(np.unique(subset_labels)) < 2:
                    auc = 0.0
                    y_pred = (subset_probs >= threshold).astype(int)
                    tp = int(np.sum((subset_labels == 1) & (y_pred == 1)))
                    tn = int(np.sum((subset_labels == 0) & (y_pred == 0)))
                    fp = int(np.sum((subset_labels == 0) & (y_pred == 1)))
                    fn = int(np.sum((subset_labels == 1) & (y_pred == 0)))
                    sen = tp / (tp + fn + 1e-16)
                    spc = tn / (tn + fp + 1e-16)
                    acc = (tp + tn) / (tp + tn + fp + fn + 1e-16)
                else:
                    metrics_result = calculate_metrics(subset_labels, subset_probs, threshold=threshold)
                    auc = metrics_result.auc
                    acc = metrics_result.accuracy
                    sen = metrics_result.sensitivity
                    spc = metrics_result.specificity

                label_map = SUBSET_VALUE_LABELS.get(indicator_name, {})
                subset_label = label_map.get(int(val), str(int(val)))

                results.append({
                    '临床指标': indicator_name,
                    '子集值': int(val),
                    '子集标签': subset_label,
                    '样本数': n_samples,
                    'AUC': auc,
                    'Acc': acc,
                    'Sen': sen,
                    'Spc': spc,
                })

        return pd.DataFrame(results)

    def analyze_prob_details(
        self,
        patient_ids: np.ndarray,
        probs: np.ndarray,
        labels: np.ndarray
    ) -> pd.DataFrame:
        """Generate per-sample probability details with subset membership."""
        if not (len(patient_ids) == len(probs) == len(labels)):
            raise ValueError("patient_ids, probs, and labels must have the same length")

        rows = []
        for indicator_name in self.resolved_indicators.keys():
            indicator_values = self.get_indicator_values(patient_ids, indicator_name)
            valid_mask = ~np.isnan(indicator_values)

            if not np.any(valid_mask):
                continue

            valid_ids = patient_ids[valid_mask]
            valid_probs = probs[valid_mask]
            valid_labels = labels[valid_mask]
            valid_values = indicator_values[valid_mask]

            label_map = SUBSET_VALUE_LABELS.get(indicator_name, {})
            for pid, prob, label, val in zip(valid_ids, valid_probs, valid_labels, valid_values):
                rows.append({
                    'ID': str(pid),
                    'Prob': float(prob),
                    'Label': int(label),
                    '临床指标': indicator_name,
                    '子集值': int(val),
                    '子集标签': label_map.get(int(val), str(int(val))),
                })

        return pd.DataFrame(rows, columns=['ID', 'Prob', 'Label', '临床指标', '子集值', '子集标签'])


# ==============================================================================
# Plotting
# ==============================================================================

def plot_feature_weights(
    coefs: np.ndarray,
    save_path: Path,
    feature_names: Optional[List[str]] = None,
    alpha: Optional[float] = None
):
    """Plot feature weight bar chart."""
    set_plot_config()
    
    if feature_names is None:
        feature_names = [f"feature_{i}" for i in range(len(coefs))]
    
    nonzero_indices = np.nonzero(coefs)[0]
    nonzero_names = [feature_names[i] for i in nonzero_indices]
    nonzero_values = coefs[coefs != 0]
    
    x_values = np.arange(len(nonzero_names))
    
    # Generate colors
    n_colors = len(nonzero_names)
    colors = []
    for i in range(n_colors // 2):
        colors.append([i / max(n_colors // 2, 1), 0.5, 0.5])
    colors = colors + colors[::-1]
    if len(colors) < n_colors:
        colors.extend([[0.5, 0.5, 0.5]] * (n_colors - len(colors)))
    
    plt.figure(figsize=(15, 16))
    plt.bar(x_values, nonzero_values, color=colors[:n_colors], alpha=0.8)
    plt.xticks(x_values, nonzero_names, rotation='vertical', ha='right', va='top')
    plt.xlabel('Feature Name', fontsize=18, fontweight='bold')
    plt.ylabel('Weight', fontsize=18, fontweight='bold')
    plt.tight_layout()
    
    filename = f"{alpha}_feature_weight.png" if alpha else "feature_weight.png"
    plt.savefig(save_path / filename, dpi=600)
    plt.close()


# ==============================================================================
# Classifier Parameter Grids
# ==============================================================================

CLASSIFIER_PARAMS = {
    'svm': {
        'kernel': ['rbf', 'sigmoid'],
        'C': [1e-8, 1e-6, 1e-5, 1e-3, 1e-2, 1e-1, 1, 32, 100, 200, 300, 400, 1000],
        'gamma': [1e-8, 1e-6, 1e-5, 1e-3, 1e-2, 1e-1],
        'class_weight': [{0: 0.55, 1: 1}, {0: 0.6, 1: 1}, {0: 0.65, 1: 1}, 
                        {0: 0.7, 1: 1}, {0: 0.8, 1: 1}, {0: 1, 1: 1}],
        'max_iter': [2, 4, 10, 20, 25, 30, 40, 60, 140, 200, 300, -1]
    },
    'lr': {
        'penalty': ['l2'],
        'C': [0.01, 0.05, 0.1, 0.3, 0.4, 0.5, 0.6, 0.8, 1.0, 1.2],
        'max_iter': [10, 20, 40, 60, 80, 100, 150, 200, 400, 500],
        'tol': [1e-5, 1e-4, 1e-3, 1e-2]
    },
    'rf': {
        'n_estimators': [101, 151, 201, 251, 301, 351, 401, 451, 501, 551, 800]
    },
    'gdbt': {
        'learning_rate': [0.1],
        'n_estimators': [100, 150, 200, 250, 300],
        'max_depth': [3, 5, 7, 9, 11],
        'subsample': [1.0]
    }
}


def get_classifier(name: str):
    """Get classifier instance by name."""
    classifiers = {
        'svm': SVC(probability=True, verbose=False),
        'lr': LogisticRegression(verbose=False),
        'rf': RandomForestClassifier(),
        'gdbt': GradientBoostingClassifier()
    }
    return classifiers.get(name)


def get_param_grid(name: str):
    """Get parameter grid for classifier."""
    return [CLASSIFIER_PARAMS.get(name, {})]



# ==============================================================================
# Main Classifier Class
# ==============================================================================

class MedicalImageClassifier:
    """Main classifier for medical image analysis."""
    
    def __init__(self, config: ClassifierConfig):
        self.config = config
        self.save_path = config.save_path
        self.feature_selector = FeatureSelector()
        
        # Data containers
        self.train_data: Dict[str, PatientData] = {}
        self.test_data: Dict[str, PatientData] = {}
        self.labels: np.ndarray = np.array([])
        self.labels_test: np.ndarray = np.array([])
    
    # --------------------------------------------------------------------------
    # Data Loading
    # --------------------------------------------------------------------------
    
    def load_patient_labels(self, excel_path: Path) -> Tuple[Dict, Dict]:
        """Load patient labels from Excel file."""
        df = pd.read_excel(excel_path)
        data_dict = df.to_dict(orient='index')
        
        train_labels = {}
        test_labels = {}
        
        for row in data_dict.values():
            patient_id = row['ID']
            if row.get('Use') == 1:
                if row.get('Type') == 'train':
                    train_labels[patient_id] = row['Label']
                else:
                    test_labels[patient_id] = row['Label']
        
        return train_labels, test_labels
    
    def load_features_from_h5(self, h5_path: Path) -> np.ndarray:
        """Load features from H5 file."""
        with h5py.File(h5_path, 'r') as f:
            features = f['f_values'][:][0]
        return features
    
    def load_patient_data(
        self, 
        patient_labels: Dict[str, int],
        radio_path: Path,
        clip_path: Path,
        is_test: bool = False
    ) -> Dict[str, PatientData]:
        """Load all patient data from H5 files."""
        data = {}
        
        for patient_id, label in patient_labels.items():
            try:
                # Load features
                if 'clip' in self.config.data_types:
                    clip_feat = self.load_features_from_h5(clip_path / f"{patient_id}.h5")
                else:
                    clip_feat = np.array([])
                
                if 'radio' in self.config.data_types:
                    radio_feat = self.load_features_from_h5(radio_path / f"{patient_id}.h5")
                else:
                    radio_feat = np.array([])
                
                # Handle single modality case
                if len(self.config.data_types) == 1:
                    if 'radio' in self.config.data_types:
                        clip_feat = radio_feat.copy()
                    else:
                        radio_feat = clip_feat.copy()
                
                data[patient_id] = PatientData(
                    patient_id=str(patient_id),
                    clip_feature=clip_feat,
                    radio_feature=radio_feat,
                    label=int(label)
                )
            except Exception as e:
                print(f"Error loading patient {patient_id}: {e}")
        
        return data
    
    def dict_to_arrays(self, data: Dict[str, PatientData]) -> Tuple:
        """Convert patient data dictionary to arrays."""
        ids = []
        labels = []
        clip_features = []
        radio_features = []
        
        for patient_id, patient in data.items():
            ids.append(patient_id)
            labels.append(patient.label)
            clip_features.append(patient.clip_feature)
            radio_features.append(patient.radio_feature)
        
        return (
            np.array(ids),
            np.array(labels),
            np.array(clip_features),
            np.array(radio_features)
        )
    
    def load_clinical_features(
        self,
        patient_ids: np.ndarray,
        excel_path: Path,
        columns: List[int]
    ) -> np.ndarray:
        """Load clinical features from Excel."""
        df = pd.read_excel(excel_path)
        values = df.values
        
        features = []
        for pid in patient_ids:
            try:
                idx = np.argwhere(values == pid)[0, 0]
                features.append(values[idx, columns])
            except IndexError:
                print(f"Patient {pid} not found in clinical data")
                features.append([0] * len(columns))
        
        return np.array(features)
    
    # --------------------------------------------------------------------------
    # Feature Processing
    # --------------------------------------------------------------------------
    
    def normalize_minmax(
        self,
        train: np.ndarray,
        test: Optional[np.ndarray] = None
    ) -> Tuple[np.ndarray, Optional[np.ndarray]]:
        """Min-max normalization."""
        train_min = train.min(axis=0)
        train_max = train.max(axis=0)
        
        train_norm = (train - train_min) / (train_max - train_min + 1e-12)
        
        if test is not None:
            test_norm = (test - train_min) / (train_max - train_min + 1e-12)
            return train_norm, test_norm
        
        return train_norm, None
    
    def concatenate_features(self, *features: np.ndarray) -> np.ndarray:
        """Concatenate multiple feature arrays."""
        return np.hstack([f for f in features if f.size > 0])
    
    def _get_feature_names(self, n_radio: int, n_clip: int) -> Tuple[List[str], List[str], Dict[str, str]]:
        """Generate feature names for radio and clip features.
        
        Returns:
            original_names: Original feature names (radio + clip)
            radiomic_aliases: Aliases with Radiomicxx for radio features, DLx for clip
            radiomic_map: Mapping from Radiomic alias to original radio name
        """
        radio_names = []
        if self.config.feature_excel_path and self.config.feature_excel_path.exists():
            try:
                df = pd.read_excel(self.config.feature_excel_path, header=None)
                values = df.iloc[:, 0].dropna().astype(str).tolist()
                if len(values) >= n_radio:
                    radio_names = values[:n_radio]
                else:
                    df2 = pd.read_excel(self.config.feature_excel_path)
                    values = [str(c) for c in df2.columns.tolist()]
                    if len(values) >= n_radio:
                        radio_names = values[:n_radio]
            except Exception as e:
                print(f"Failed to load feature names from excel: {e}")
        
        if len(radio_names) < n_radio:
            radio_names = [f"radio_feature_{i}" for i in range(n_radio)]
        
        # Create radiomic aliases for SHAP plotting
        radiomic_aliases = [f"Radiomic{i+1}" for i in range(n_radio)]
        radiomic_map = {radiomic_aliases[i]: radio_names[i] for i in range(n_radio)}
        
        clip_names = [f"DL{i}" for i in range(n_clip)]
        return radio_names + clip_names, radiomic_aliases + clip_names, radiomic_map
    
    def _save_selected_features(
        self,
        train_features: np.ndarray,
        test_features: Optional[np.ndarray],
        train_ids: np.ndarray,
        test_ids: Optional[np.ndarray],
        train_labels: np.ndarray,
        test_labels: Optional[np.ndarray],
        feature_names: List[str],
        save_path: Path
    ):
        """Save selected feature values to Excel."""
        try:
            df_train = pd.DataFrame(train_features, columns=feature_names)
            df_train.insert(0, 'ID', train_ids)
            df_train['Label'] = train_labels
            
            save_file = save_path / 'AllFeatureFinal.xlsx'
            if test_features is not None and test_ids is not None and test_labels is not None:
                df_test = pd.DataFrame(test_features, columns=feature_names)
                df_test.insert(0, 'ID', test_ids)
                df_test['Label'] = test_labels
                
                with pd.ExcelWriter(save_file) as writer:
                    df_train.to_excel(writer, sheet_name='AllFeatureFinal', index=False)
                    df_test.to_excel(writer, sheet_name='AllFeatureOutFinal', index=False)
            else:
                df_train.to_excel(save_file, sheet_name='AllFeatureFinal', index=False)
            
            print(f"Selected features saved to {save_file}")
        except Exception as e:
            print(f"Error saving selected features: {e}")
    
    # --------------------------------------------------------------------------
    # Cross-Validation
    # --------------------------------------------------------------------------
    
    def run_cross_validation(
        self,
        features: np.ndarray,
        labels: np.ndarray,
        features_test: Optional[np.ndarray],
        labels_test: Optional[np.ndarray],
        clinical_train: np.ndarray,
        clinical_test: np.ndarray,
        patient_ids: np.ndarray,
        patient_ids_test: np.ndarray,
        task: str,
        save_path: Path,
        feature_names: Optional[List[str]] = None,
        radiomic_aliases: Optional[List[str]] = None,
        subset_analyzer: Optional[ClinicalSubsetAnalyzer] = None
    ) -> Dict:
        """Run stratified k-fold cross-validation."""
        results = {
            'train_probs': [],
            'train_labels': [],
            'train_ids': [],
            'val_probs': [],
            'test_probs': [],
            'val_labels': [],
            'val_ids': []
        }
        
        skf = StratifiedKFold(
            n_splits=self.config.k_fold,
            random_state=self.config.cv_random_state,
            shuffle=True
        )
        
        fold = 1
        for train_idx, val_idx in skf.split(features, labels):
            print(f"Fold {fold}/{self.config.k_fold}")
            
            # Split data
            X_train, X_val = features[train_idx], features[val_idx]
            y_train, y_val = labels[train_idx], labels[val_idx]
            clin_train, clin_val = clinical_train[train_idx], clinical_train[val_idx]
            
            # Concatenate with clinical features
            X_train = self.concatenate_features(X_train, clin_train)
            X_val = self.concatenate_features(X_val, clin_val)
            
            if features_test is not None:
                X_test = self.concatenate_features(features_test, clinical_test)
            
            # Normalize
            scaler = preprocessing.StandardScaler()
            X_train = scaler.fit_transform(X_train)
            X_val = scaler.transform(X_val)
            if features_test is not None:
                X_test = scaler.transform(X_test)
            
            # Train classifier
            if self.config.use_lr:
                clf = GridSearchCV(
                    LogisticRegression(verbose=False),
                    get_param_grid('lr'),
                    scoring='roc_auc'
                )
                clf.fit(X_train, y_train)
                
                # SHAP analysis
                try:
                    # Use radiomic aliases (pre-generated before filtering) for SHAP plots only
                    # Exclude clinical features from SHAP plots
                    n_clinical = clinical_train.shape[1] if clinical_train.ndim > 1 else 1
                    n_shap_features = X_train.shape[1] - n_clinical
                    X_train_shap = X_train[:, :n_shap_features]
                    
                    all_feature_names = list(radiomic_aliases) if radiomic_aliases else (list(feature_names) if feature_names else [f"feature_{i}" for i in range(n_shap_features)])
                    
                    try:
                        explainer = shap.LinearExplainer(clf.best_estimator_, X_train)
                        shap_values = explainer.shap_values(X_train)
                    except Exception:
                        background = shap.sample(X_train, min(50, len(X_train)))
                        explainer = shap.KernelExplainer(lambda x: clf.best_estimator_.predict_proba(x)[:, 1], background)
                        shap_values = explainer.shap_values(X_train)
                    
                    if isinstance(shap_values, list):
                        shap_values_plot = shap_values[1]
                    else:
                        shap_values_plot = shap_values
                    
                    shap_values_shap = shap_values_plot[:, :n_shap_features]
                    
                    plt.figure(figsize=(12, 8))
                    shap.summary_plot(shap_values_shap, X_train_shap, feature_names=all_feature_names, show=False)
                    plt.tight_layout()
                    plt.savefig(save_path / f'shap_summary_fold{fold}.png', dpi=600, bbox_inches='tight')
                    plt.close()
                    
                    plt.figure(figsize=(12, 8))
                    shap.summary_plot(shap_values_shap, X_train_shap, feature_names=all_feature_names, plot_type="bar", show=False)
                    plt.tight_layout()
                    plt.savefig(save_path / f'shap_bar_fold{fold}.png', dpi=600, bbox_inches='tight')
                    plt.close()
                    print(f"SHAP plots saved for fold {fold}")
                except Exception as e:
                    print(f"SHAP analysis failed for fold {fold}: {e}")
                
                # Predictions
                train_prob = clf.predict_proba(X_train)[:, 1]
                val_prob = clf.predict_proba(X_val)[:, 1]
                
                # Save fold results
                self._save_fold_results(
                    save_path, fold, train_idx, train_prob, y_train,
                    val_idx, val_prob, y_val, patient_ids
                )
                
                results['train_probs'].extend(train_prob)
                results['train_labels'].extend(y_train)
                results['train_ids'].extend(patient_ids[train_idx])
                results['val_probs'].extend(val_prob)
                results['val_labels'].extend(y_val)
                results['val_ids'].extend(patient_ids[val_idx])
                
                if features_test is not None:
                    test_prob = clf.predict_proba(X_test)[:, 1]
                    results['test_probs'].append(test_prob)
                    
                    # Save per-fold test predictions
                    df_test = pd.DataFrame({
                        'ID': patient_ids_test,
                        'Prob': test_prob,
                        'Label': labels_test
                    })
                    df_test.to_excel(save_path / f"test_{fold}.xlsx", index=False)
                else:
                    test_prob = None
                
                # Save per-fold clinical subset analysis
                self._save_fold_subset_results(
                    save_path, fold, subset_analyzer,
                    patient_ids[train_idx], train_prob, y_train,
                    patient_ids[val_idx], val_prob, y_val,
                    patient_ids_test if features_test is not None else None,
                    test_prob,
                    labels_test if features_test is not None else None
                )
            
            fold += 1
        
        return results
    
    def _save_fold_results(
        self,
        save_path: Path,
        fold: int,
        train_idx: np.ndarray,
        train_prob: np.ndarray,
        train_labels: np.ndarray,
        val_idx: np.ndarray,
        val_prob: np.ndarray,
        val_labels: np.ndarray,
        patient_ids: np.ndarray
    ):
        """Save results for a single fold."""
        # Training results
        df_train = pd.DataFrame({
            'ID': patient_ids[train_idx],
            'Prob': train_prob,
            'Label': train_labels
        })
        df_train.to_excel(save_path / f"train_{fold}.xlsx", index=False)
        
        # Validation results
        df_val = pd.DataFrame({
            'ID': patient_ids[val_idx],
            'Prob': val_prob,
            'Label': val_labels
        })
        df_val.to_excel(save_path / f"valid_{fold}.xlsx", index=False)
    
    def _save_fold_subset_results(
        self,
        save_path: Path,
        fold: int,
        subset_analyzer: Optional[ClinicalSubsetAnalyzer],
        train_ids: np.ndarray,
        train_probs: np.ndarray,
        train_labels: np.ndarray,
        val_ids: np.ndarray,
        val_probs: np.ndarray,
        val_labels: np.ndarray,
        test_ids: Optional[np.ndarray] = None,
        test_probs: Optional[np.ndarray] = None,
        test_labels: Optional[np.ndarray] = None
    ):
        """Save clinical subset analysis for a single fold."""
        if subset_analyzer is None:
            return
        
        try:
            # Use fold-specific validation threshold for subset metrics
            if len(np.unique(val_labels)) >= 2:
                threshold = calculate_metrics(val_labels, val_probs).best_threshold
            else:
                threshold = 0.5
            
            subset_path = save_path / f"clinical_subset_{fold}.xlsx"
            with pd.ExcelWriter(subset_path) as writer:
                # Subset metric sheets
                train_df = subset_analyzer.analyze(
                    train_ids, train_probs, train_labels, threshold=threshold
                )
                val_df = subset_analyzer.analyze(
                    val_ids, val_probs, val_labels, threshold=threshold
                )
                train_df.to_excel(writer, sheet_name='Train', index=False)
                val_df.to_excel(writer, sheet_name='Val', index=False)
                
                # Probability detail sheets
                train_detail_df = subset_analyzer.analyze_prob_details(
                    train_ids, train_probs, train_labels
                )
                val_detail_df = subset_analyzer.analyze_prob_details(
                    val_ids, val_probs, val_labels
                )
                train_detail_df.to_excel(writer, sheet_name='Prob_Detail_Train', index=False)
                val_detail_df.to_excel(writer, sheet_name='Prob_Detail_Val', index=False)
                
                # Optional test subset
                if test_ids is not None and test_probs is not None and test_labels is not None:
                    test_df = subset_analyzer.analyze(
                        test_ids, test_probs, test_labels, threshold=threshold
                    )
                    test_df.to_excel(writer, sheet_name='Test', index=False)
                    test_detail_df = subset_analyzer.analyze_prob_details(
                        test_ids, test_probs, test_labels
                    )
                    test_detail_df.to_excel(writer, sheet_name='Prob_Detail_Test', index=False)
            
            print(f"Clinical subset fold results saved to {subset_path}")
        except Exception as e:
            print(f"Error saving clinical subset fold results for fold {fold}: {e}")
    
    # --------------------------------------------------------------------------
    # Result Saving
    # --------------------------------------------------------------------------
    
    def save_metrics(
        self,
        save_path: Path,
        metrics_in: MetricsResult,
        metrics_out: Optional[MetricsResult],
        params: Dict
    ):
        """Save metrics to Excel."""
        filename = save_path / 'all_results.xlsx'
        
        # Header
        headers = ['Round', 'Params', 'Distill_Mode', 'CV_State', 'Grid_State', 'K_Fold']
        for i, h in enumerate(headers, 1):
            write_to_excel(filename, h, 1, i)
        
        # Values
        write_to_excel(filename, '1', 2, 1)
        write_to_excel(filename, str(params), 2, 2)
        write_to_excel(filename, self.config.distill_mode, 2, 3)
        write_to_excel(filename, str(self.config.cv_random_state), 2, 4)
        write_to_excel(filename, str(self.config.grid_random_state), 2, 5)
        write_to_excel(filename, str(self.config.k_fold), 2, 6)
        
        # LR metrics
        if self.config.use_lr:
            write_to_excel(filename, 'LR', 1, 7)
            write_to_excel(filename, 'In_AUC', 1, 8)
            write_to_excel(filename, metrics_in.auc, 2, 8)
            write_to_excel(filename, 'In_Acc', 1, 9)
            write_to_excel(filename, metrics_in.accuracy, 2, 9)
            write_to_excel(filename, 'In_Sen', 1, 10)
            write_to_excel(filename, metrics_in.sensitivity, 2, 10)
            write_to_excel(filename, 'In_Spc', 1, 11)
            write_to_excel(filename, metrics_in.specificity, 2, 11)
            
            if metrics_out:
                write_to_excel(filename, 'Out_AUC', 1, 13)
                write_to_excel(filename, metrics_out.auc, 2, 13)
    
    def save_probabilities(
        self,
        save_path: Path,
        probs: List,
        labels: List,
        ids: List,
        round_num: int = 1
    ):
        """Save probability scores to Excel."""
        filename = save_path / 'In_prob_results.xlsx'
        sheet = 'LR'
        
        col_start = round_num * 4 - 3
        write_to_excel(filename, f'Round_{round_num}', 1, col_start, sheet)
        write_to_excel(filename, 'ID', 1, col_start + 1, sheet)
        write_to_excel(filename, 'Prob', 1, col_start + 2, sheet)
        write_to_excel(filename, 'Label', 1, col_start + 3, sheet)
        
        for i, (pid, prob, label) in enumerate(zip(ids, probs, labels)):
            write_to_excel(filename, str(pid), i + 2, col_start + 1, sheet)
            write_to_excel(filename, prob, i + 2, col_start + 2, sheet)
            write_to_excel(filename, label, i + 2, col_start + 3, sheet)
    
    # --------------------------------------------------------------------------
    # Main Pipeline
    # --------------------------------------------------------------------------
    
    def run(self):
        """Execute full classification pipeline."""
        random.seed(6666)
        subset_analyzer = ClinicalSubsetAnalyzer(self.config.ser_path)

        # Load labels
        train_labels, test_labels = self.load_patient_labels(self.config.ser_path)
        print(f"Train samples: {len(train_labels)}, Test samples: {len(test_labels)}")

        # Create output directory
        ensure_dir(self.save_path)

        # Load training data for both tasks
        # Store data for each task to combine later
        task_results = {}

        for task in self.config.tasks:
            print(f"\n{'='*50}")
            print(f"Processing Task: {task}")
            print(f"{'='*50}")
            
            radio_path = getattr(self.config, f'radio_feature_path_{task}')
            clip_path = getattr(self.config, f'clip_feature_path_{task}')
            
            # Load data for current task
            train_data = self.load_patient_data(
                train_labels, radio_path, clip_path
            )
            test_data = None
            if self.config.use_external_test:
                test_data = self.load_patient_data(
                    test_labels, radio_path, clip_path, is_test=True
                )
            
            # Convert to arrays
            ids, labels, clip_feats, radio_feats = self.dict_to_arrays(train_data)
            self.labels = labels
            
            ids_test, labels_test, clip_test, radio_test = None, None, None, None
            if self.config.use_external_test and test_data:
                ids_test, labels_test, clip_test, radio_test = self.dict_to_arrays(test_data)
                self.labels_test = labels_test
            
            # Load clinical and SER features
            clinical_cols = [7, 8, 9, 10]  # ER, PR, HER2, Ki67
            clinical_train = self.load_clinical_features(
                ids, self.config.ser_path, clinical_cols
            )
            
            ser_cols = [1] if task == '0' else [5]
            ser_train = self.load_clinical_features(
                ids, self.config.ser_path, ser_cols
            )
            
            if self.config.use_external_test:
                clinical_test = self.load_clinical_features(
                    ids_test, self.config.ser_path, clinical_cols
                )
                ser_test = self.load_clinical_features(
                    ids_test, self.config.ser_path, ser_cols
                )
            else:
                clinical_test = np.array([])
                ser_test = np.array([])
            
            # Normalize features
            clip_test_norm = clip_test if clip_test is not None else None
            radio_test_norm = radio_test if radio_test is not None else None
            
            clip_feats, clip_test_norm = self.normalize_minmax(clip_feats, clip_test_norm)
            radio_feats, radio_test_norm = self.normalize_minmax(radio_feats, radio_test_norm)
            
            ser_test_norm = ser_test if ser_test.size > 0 else None
            ser_train_norm, ser_test_norm = self.normalize_minmax(ser_train, ser_test_norm)
            
            # Generate feature names (original + radiomic aliases for SHAP)
            n_radio_orig = radio_feats.shape[1] if radio_feats.ndim > 1 else 0
            n_clip_orig = clip_feats.shape[1] if clip_feats.ndim > 1 else 0
            feature_names, radiomic_aliases, radiomic_map = self._get_feature_names(n_radio_orig, n_clip_orig)
            
            # Feature selection with SelectKBest - separate selectors for each modality
            selector_radio = FeatureSelector()
            selector_clip = FeatureSelector()
            
            radio_feats, _ = selector_radio.select_k_best(
                radio_feats, labels, 
                self.config.select_kbest_k, 
                self.config.select_kbest_func
            )
            clip_feats, _ = selector_clip.select_k_best(
                clip_feats, labels,
                self.config.select_kbest_k,
                self.config.select_kbest_func
            )
            
            if self.config.use_external_test and clip_test_norm is not None:
                # Apply same selection to test using respective masks
                radio_test_norm = radio_test_norm[:, selector_radio.selected_indices] if radio_test_norm is not None else None
                clip_test_norm = clip_test_norm[:, selector_clip.selected_indices] if clip_test_norm is not None else None
            
            # Update feature names after SelectKBest
            if feature_names:
                radio_mask = selector_radio.selected_indices if selector_radio.selected_indices is not None else np.ones(n_radio_orig, dtype=bool)
                clip_mask = selector_clip.selected_indices if selector_clip.selected_indices is not None else np.ones(n_clip_orig, dtype=bool)
                feature_names = [feature_names[i] for i in range(n_radio_orig) if radio_mask[i]] + \
                               [feature_names[i] for i in range(n_radio_orig, n_radio_orig + n_clip_orig) if clip_mask[i - n_radio_orig]]
                radiomic_aliases = [radiomic_aliases[i] for i in range(n_radio_orig) if radio_mask[i]] + \
                                    [radiomic_aliases[i] for i in range(n_radio_orig, n_radio_orig + n_clip_orig) if clip_mask[i - n_radio_orig]]
            
            # Concatenate features
            all_features = self.concatenate_features(radio_feats, clip_feats)
            all_test = self.concatenate_features(radio_test_norm, clip_test_norm) if self.config.use_external_test else None
            
            # LASSO feature selection
            lasso_path = self.save_path / f'task_{task}' / 'lassocv'
            ensure_dir(lasso_path)
            
            if self.config.use_lassocv:
                coef, intercept = self.feature_selector.lasso_cv_select(
                    all_features, labels, lasso_path, self.config.lassocv_k,
                    feature_names=feature_names
                )
                selected_mask = coef != 0
                all_features = all_features[:, selected_mask]
                if all_test is not None:
                    all_test = all_test[:, selected_mask]
                if feature_names:
                    feature_names = [feature_names[i] for i in range(len(feature_names)) if selected_mask[i]]
                if radiomic_aliases:
                    radiomic_aliases = [radiomic_aliases[i] for i in range(len(radiomic_aliases)) if selected_mask[i]]
            
            # Append SER feature after LASSO selection (always included)
            if ser_train_norm is not None and ser_train_norm.size > 0:
                if ser_train_norm.ndim == 1:
                    ser_train_norm = ser_train_norm.reshape(-1, 1)
                all_features = self.concatenate_features(all_features, ser_train_norm)
                if all_test is not None and ser_test_norm is not None and ser_test_norm.size > 0:
                    if ser_test_norm.ndim == 1:
                        ser_test_norm = ser_test_norm.reshape(-1, 1)
                    all_test = self.concatenate_features(all_test, ser_test_norm)
                if feature_names is not None:
                    feature_names = feature_names + ['SER']
                if radiomic_aliases is not None:
                    radiomic_aliases = radiomic_aliases + ['SER']
            
            # Save selected feature values to Excel
            self._save_selected_features(
                all_features, all_test,
                ids, ids_test if self.config.use_external_test else None,
                labels, labels_test if self.config.use_external_test else None,
                feature_names, lasso_path
            )
            
            # Run cross-validation
            task_save_path = self.save_path / f'task_{task}'
            task_save_path.mkdir(parents=True, exist_ok=True)
            
            # Save pre-filtering radiomic name mapping table (once per task)
            if radiomic_map:
                df_map = pd.DataFrame(list(radiomic_map.items()), columns=['SHAP_Name', 'Original_Name'])
                df_map.to_excel(task_save_path / 'radiomic_name_mapping.xlsx', index=False)
                print(f"Radiomic name mapping saved for task {task}")
            
            # Save post-filtering radiomic name mapping table (only selected features)
            if radiomic_aliases and feature_names:
                selected_mapping = []
                for alias, orig in zip(radiomic_aliases, feature_names):
                    if not alias.startswith('DL'):
                        selected_mapping.append({'SHAP_Name': alias, 'Original_Name': orig})
                if selected_mapping:
                    df_map_selected = pd.DataFrame(selected_mapping)
                    df_map_selected.to_excel(task_save_path / 'radiomic_name_mapping_selected.xlsx', index=False)
                    print(f"Selected radiomic name mapping saved for task {task}")
            
            results = self.run_cross_validation(
                all_features, labels,
                all_test, self.labels_test if self.config.use_external_test else None,
                clinical_train, clinical_test,
                ids, ids_test if self.config.use_external_test else np.array([]),
                task, task_save_path,
                feature_names=feature_names,
                radiomic_aliases=radiomic_aliases,
                subset_analyzer=subset_analyzer
            )
            
            # Calculate and save metrics
            metrics_in = calculate_metrics(
                np.array(results['val_labels']),
                np.array(results['val_probs'])
            )
            
            metrics_out = None
            test_probs_mean = None
            if self.config.use_external_test and results['test_probs']:
                test_probs_mean = np.mean(results['test_probs'], axis=0)
                metrics_out = calculate_metrics(
                    self.labels_test, test_probs_mean,
                    threshold=metrics_in.best_threshold
                )
            
            self.save_metrics(
                task_save_path, metrics_in, metrics_out,
                {'lassocv_k': self.config.lassocv_k}
            )
            
            self.save_probabilities(
                task_save_path,
                results['val_probs'],
                results['val_labels'],
                results['val_ids']
            )

            # Aggregate per-fold train predictions so each patient appears once
            train_ids_agg, train_probs_agg, train_labels_agg = aggregate_fold_predictions(
                np.array(results['train_ids']),
                np.array(results['train_probs']),
                np.array(results['train_labels'])
            )

            # Save clinical subset metrics
            try:
                train_df = subset_analyzer.analyze(
                    train_ids_agg, train_probs_agg, train_labels_agg,
                    threshold=metrics_in.best_threshold
                )
                val_df = subset_analyzer.analyze(
                    np.array(results['val_ids']),
                    np.array(results['val_probs']),
                    np.array(results['val_labels']),
                    threshold=metrics_in.best_threshold
                )

                # Probability detail tables
                train_detail_df = subset_analyzer.analyze_prob_details(
                    train_ids_agg, train_probs_agg, train_labels_agg
                )
                val_detail_df = subset_analyzer.analyze_prob_details(
                    np.array(results['val_ids']),
                    np.array(results['val_probs']),
                    np.array(results['val_labels'])
                )

                subset_path = task_save_path / 'clinical_subset_metrics.xlsx'
                with pd.ExcelWriter(subset_path) as writer:
                    train_df.to_excel(writer, sheet_name='Train', index=False)
                    val_df.to_excel(writer, sheet_name='Val', index=False)
                    train_detail_df.to_excel(writer, sheet_name='Prob_Detail_Train', index=False)
                    val_detail_df.to_excel(writer, sheet_name='Prob_Detail_Val', index=False)

                    if self.config.use_external_test and results['test_probs']:
                        test_df = subset_analyzer.analyze(
                            ids_test,
                            test_probs_mean,
                            self.labels_test,
                            threshold=metrics_in.best_threshold
                        )
                        test_df.to_excel(writer, sheet_name='Test', index=False)
                        test_detail_df = subset_analyzer.analyze_prob_details(
                            ids_test, test_probs_mean, self.labels_test
                        )
                        test_detail_df.to_excel(writer, sheet_name='Prob_Detail_Test', index=False)

                print(f"Clinical subset metrics saved to {subset_path}")
            except Exception as e:
                print(f"Error saving clinical subset metrics: {e}")

            # Print results for current task
            print(f"\n{'='*50}")
            print(f"Task {task} Results:")
            print(f"{'='*50}")
            print(f"Validation - AUC: {metrics_in.auc:.4f}, "
                  f"Acc: {metrics_in.accuracy:.4f}, "
                  f"Sen: {metrics_in.sensitivity:.4f}, "
                  f"Spc: {metrics_in.specificity:.4f}")
            if metrics_out:
                print(f"Test       - AUC: {metrics_out.auc:.4f}, "
                      f"Acc: {metrics_out.accuracy:.4f}, "
                      f"Sen: {metrics_out.sensitivity:.4f}, "
                      f"Spc: {metrics_out.specificity:.4f}")
            print(f"{'='*50}\n")

            # Store results for combining later
            task_results[task] = {
                'train_probs': train_probs_agg,
                'train_labels': train_labels_agg,
                'train_ids': train_ids_agg,
                'val_probs': np.array(results['val_probs']),
                'val_labels': np.array(results['val_labels']),
                'val_ids': np.array(results['val_ids']),
                'val_ser': ser_train_norm.copy(),
                'test_probs': test_probs_mean if results['test_probs'] else None,
                'test_labels': self.labels_test if self.config.use_external_test else None,
                'test_ids': ids_test.copy() if self.config.use_external_test and ids_test is not None else None,
                'test_ser': ser_test_norm.copy() if ser_test_norm is not None and ser_test_norm.size > 0 else None
            }

            gc.collect()
            print(f"Task {task} completed!")
        
        # Combine results from all tasks (average of task 0 and task 6)
        if len(self.config.tasks) == 2 and '0' in task_results and '6' in task_results:
            self._save_combined_results(task_results, subset_analyzer)
    
    def _save_combined_results(self, task_results: Dict, subset_analyzer: ClinicalSubsetAnalyzer):
        """Save combined results using equal-weight fusion (w0=0.5, w6=0.5)."""
        print("\n" + "="*50)
        print("Combining Task 0 and Task 6 Results (Equal Weight w0=0.5, w6=0.5)")
        print("="*50)
        
        # Get validation data
        prob_0 = np.asarray(task_results['0']['val_probs'], dtype=np.float64)
        prob_6 = np.asarray(task_results['6']['val_probs'], dtype=np.float64)
        labels_val = task_results['0']['val_labels']
        ids_val = task_results['0']['val_ids']
        
        # Baseline
        auc_task0_only = metrics.roc_auc_score(labels_val, prob_0)
        auc_task6_only = metrics.roc_auc_score(labels_val, prob_6)
        print(f"  [Baseline] Task0-only Val_AUC: {auc_task0_only:.4f}")
        print(f"  [Baseline] Task6-only Val_AUC: {auc_task6_only:.4f}")
        
        # Equal fusion: w0=0.5, w6=0.5
        W6 = 0.50
        W0 = 0.50
        prob_combined = W0 * prob_0 + W6 * prob_6
        diff = np.abs(prob_0 - prob_6)
        
        metrics_combined = calculate_metrics(labels_val, prob_combined)
        print(f"  Equal Fusion (w0={W0:.2f}, w6={W6:.2f}) Val_AUC: {metrics_combined.auc:.4f}, "
              f"Acc: {metrics_combined.accuracy:.4f}, "
              f"Sen: {metrics_combined.sensitivity:.4f}, "
              f"Spc: {metrics_combined.specificity:.4f}")

        # Combined training data
        train_prob_0 = np.asarray(task_results['0']['train_probs'], dtype=np.float64)
        train_prob_6 = np.asarray(task_results['6']['train_probs'], dtype=np.float64)
        labels_train = task_results['0']['train_labels']
        ids_train = task_results['0']['train_ids']

        prob_train_combined = W0 * train_prob_0 + W6 * train_prob_6
        metrics_train_combined = calculate_metrics(labels_train, prob_train_combined)
        print(f"  Equal Fusion Train - AUC: {metrics_train_combined.auc:.4f}, "
              f"Acc: {metrics_train_combined.accuracy:.4f}, "
              f"Sen: {metrics_train_combined.sensitivity:.4f}, "
              f"Spc: {metrics_train_combined.specificity:.4f}")

        # Save combined results
        combined_path = self.save_path / 'combined_0_6_optimized'
        ensure_dir(combined_path)
        
        df_fusion = pd.DataFrame({
            'ID': ids_val,
            'Prob_Task0': prob_0,
            'Prob_Task6': prob_6,
            'Prob_Combined': prob_combined,
            'Prob_Diff': diff,
            'Weight_Task0': W0,
            'Weight_Task6': W6,
            'Label': labels_val
        })
        df_fusion.to_excel(combined_path / 'fusion_details.xlsx', index=False)
        
        # Pre-compute combined test metrics if both tasks have test data
        prob_test_0 = None
        prob_test_6 = None
        labels_test = None
        ids_test = None
        prob_test_combined = None
        diff_test = None
        metrics_test_combined = None
        if task_results['0']['test_probs'] is not None and task_results['6']['test_probs'] is not None:
            prob_test_0 = np.asarray(task_results['0']['test_probs'], dtype=np.float64)
            prob_test_6 = np.asarray(task_results['6']['test_probs'], dtype=np.float64)
            labels_test = task_results['0']['test_labels']
            ids_test = task_results['0']['test_ids']
            
            prob_test_combined = W0 * prob_test_0 + W6 * prob_test_6
            diff_test = np.abs(prob_test_0 - prob_test_6)
            
            metrics_test_combined = calculate_metrics(labels_test, prob_test_combined)
        
        try:
            combined_train_df = subset_analyzer.analyze(
                ids_train, prob_train_combined, labels_train,
                threshold=metrics_train_combined.best_threshold
            )
            combined_val_df = subset_analyzer.analyze(
                ids_val, prob_combined, labels_val,
                threshold=metrics_combined.best_threshold
            )

            subset_path = combined_path / 'clinical_subset_metrics.xlsx'
            with pd.ExcelWriter(subset_path) as writer:
                combined_train_df.to_excel(writer, sheet_name='Combined_Train', index=False)
                combined_val_df.to_excel(writer, sheet_name='Combined_Val', index=False)

                # Detail sheets
                train_detail_df = subset_analyzer.analyze_prob_details(
                    ids_train, prob_train_combined, labels_train
                )
                val_detail_df = subset_analyzer.analyze_prob_details(
                    ids_val, prob_combined, labels_val
                )
                train_detail_df.to_excel(writer, sheet_name='Prob_Detail_Combined_Train', index=False)
                val_detail_df.to_excel(writer, sheet_name='Prob_Detail_Combined_Val', index=False)

                if task_results['0']['test_probs'] is not None and task_results['6']['test_probs'] is not None:
                    combined_test_df = subset_analyzer.analyze(
                        ids_test, prob_test_combined, labels_test,
                        threshold=metrics_test_combined.best_threshold
                    )
                    combined_test_df.to_excel(writer, sheet_name='Combined_Test', index=False)
                    test_detail_df = subset_analyzer.analyze_prob_details(
                        ids_test, prob_test_combined, labels_test
                    )
                    test_detail_df.to_excel(writer, sheet_name='Prob_Detail_Combined_Test', index=False)

            print(f"Combined clinical subset metrics saved to {subset_path}")
        except Exception as e:
            print(f"Error saving combined clinical subset metrics: {e}")
        
        self.save_metrics(
            combined_path, metrics_combined, metrics_test_combined,
            {'method': 'equal_w0.5_w6.0.5', 'params': str({'w0': W0, 'w6': W6})}
        )
        
        # Test results
        if task_results['0']['test_probs'] is not None and task_results['6']['test_probs'] is not None:
            print(f"Equal Fusion Test           - AUC: {metrics_test_combined.auc:.4f}, "
                  f"Acc: {metrics_test_combined.accuracy:.4f}, "
                  f"Sen: {metrics_test_combined.sensitivity:.4f}, "
                  f"Spc: {metrics_test_combined.specificity:.4f}")
            
            df_fusion_test = pd.DataFrame({
                'ID': ids_test,
                'Prob_Task0': prob_test_0,
                'Prob_Task6': prob_test_6,
                'Prob_Combined': prob_test_combined,
                'Prob_Diff': diff_test,
                'Label': labels_test
            })
            df_fusion_test.to_excel(combined_path / 'fusion_details_test.xlsx', index=False)


# ==============================================================================
# Main Entry Point
# ==============================================================================

def main():
    """Main entry point."""
    current_time = datetime.datetime.now()
    time_str = current_time.strftime("%m-%d-%H-%M")
    # NOW: 112, 2024, 456, 101112, 102, 223, 589, 775, 911, 101
    seeds = [112, 2024, 456, 101112, 102, 223, 589, 775, 911, 101]# 112, 2024, 456, 101112, 131415, 223, 589, 775, 886, 911, 1011
    base_save_path = Path(r'D:\1实验室项目\DGbreast\Result\ALL_MultiSeed_subset') / time_str # 1234,  1422,  1777, 1991, 2026
    
    for seed in seeds:
        print(f"\n{'='*60}")
        print(f"Running with seed: {seed}")
        print(f"{'='*60}")
        
        # Configuration
        config = ClassifierConfig(
            save_path=base_save_path / f'seed_{seed}',
            radio_feature_path_0=Path(r'D:\1实验室项目\DGbreast\Radio_feature_zscore\qian'),
            clip_feature_path_0=Path(r'D:\1实验室项目\DGbreast\Deep_Feature\qian'),
            radio_feature_path_6=Path(r'D:\1实验室项目\DGbreast\Radio_feature_zscore\hou'),
            clip_feature_path_6=Path(r'D:\1实验室项目\DGbreast\Deep_FeatureA\hou'),
            ser_path=Path(r"D:\1实验室项目\DGbreast\NEW\excel\original\ALL_Result_Final_TRY_Correct-Complete-filled_GZ-updated.xlsx"),
            feature_excel_path=Path(r"D:\SY-NAC-xiudinghou\feature_title.xlsx"),
            tasks=['0', '6'],
            data_types=['clip', 'radio'],
            use_external_test=True,
            k_fold=5,
            cv_random_state=seed,
            grid_random_state=seed,
            use_lr=True,
            fusion_min_improvement=0.0  # allow fusion even if only matching Task6-only AUC
        )
        
        print(f"Output path: {config.save_path}")
        
        # Run classifier
        classifier = MedicalImageClassifier(config)
        classifier.run()
    
    print(f"\n{'='*60}")
    print("All seeds completed!")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
