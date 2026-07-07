import os
from radiomics import featureextractor as FEE
from openpyxl import Workbook
from openpyxl import load_workbook
import shutil

from scipy.ndimage import zoom
# from wama.utils import *
import pandas as pd
import SimpleITK as sitk
import numpy as np
import xlrd
import xlwt
import xlutils
import os
sep = os.sep

# 保存实验信息到txt文件中
def save_exinfor_txt(input, savepath):
    sep = os.sep
    filename = savepath + sep + '实验信息.txt'
    file = open(filename, 'a+')
    file.write(input)
    file.close()

# 存入数据到excel中,一个个存
def writesimple2(path, data, index, column, sheetname='Sheet'):  # index是行数,colum是列数
    '''
    :param path: 存储地址
    :param list: 数据
    :param index: 行
    :param column: 列
    :param sheetname: sheet名称,默认为Sheet
    :return:
    '''
    if os.path.exists(path):
        bg = xlrd.open_workbook(path)
        sheets = bg.sheet_names() # 获取所有sheet名字
        if sheetname in sheets: # 存在该sheet,则追加数据
            Index = sheets.index(sheetname) # sheet对应的索引
            bg1 = xlutils.copy(bg) # xlrd转xlwt
            writesheet = bg1.get_sheet(Index)
            writesheet.write(index, column, str(data))
            bg1.save(path)
        else:# 不存在sheet,就重新创建sheet
            bg2 = xlutils.copy(bg)  # xlrd转xlwt
            writesheet = bg2.add_sheet(sheetname)
            writesheet.write(index, column, str(data))
            bg2.save(path)
    else:
        bg = xlwt.Workbook()  # 创建一个.xlsx文件,默认生成一个名为Sheet的sheet
        sheet = bg.add_sheet(sheetname)
        sheet.write(index, column, str(data))
        bg.save(path)

# 存入数据到excel中,一个个存
def writesimple(path, data, index, column, sheetname='Sheet'):  # index是行数,colum是列数
    '''
    :param path: 存储地址
    :param list: 数据
    :param index: 行
    :param column: 列
    :param sheetname: sheet名称,默认为Sheet
    :return:
    '''
    if os.path.exists(path):
        bg = load_workbook(path)
        sheets = bg.sheetnames
        if sheetname in sheets:
            sheet = bg[sheetname]
            sheet.cell(index, column, data)
            bg.save(path)
            bg.close()
        else:
            sheet = bg.create_sheet(sheetname)
            # sheet = bg[str(sheetname)]
            sheet.cell(index, column, data)
            bg.save(path)
            bg.close()
    else:
        bg = Workbook()  # 创建一个.xlsx文件,默认生成一个名为Sheet的sheet
        # 修改默认Sheet名为自定义sheet名
        bg1 = bg['Sheet']
        bg1.title = sheetname
        sheet = bg[sheetname]
        sheet.cell(index, column, data)
        bg.save(path)
        bg.close()

# 存入数据到excel中,list按列写入
def writelist(path,list,head=None,columns=None,sheet_name = 'Sheet'):
    pf = pd.DataFrame()
    if head != None:
        pf[head] = list
        header = True
    else:
        header = False
    # index 为FALSE 表示没有第一列数字为索引
    pf.to_excel(path,encoding='utf-8', index=False, header=header, sheet_name = sheet_name,columns=columns)

# 判断文件夹是否存在
def judgedir(path, RemoveFlag=False):
    if os.path.exists(path):
        if RemoveFlag: # 该目录已经存在的话，就删除
                shutil.rmtree(path)  # 空目录,有内容的目录都可以删
                os.makedirs(path)
        else:
            pass
    else:
        os.makedirs(path)

# mask 膨胀
def mask_pengzhang(path,savepath):
    img = sitk.ReadImage(path)

    spacing = img.GetSpacing()
    origin = img.GetOrigin()
    transfmat = img.GetDirection()
    img = sitk.GetArrayFromImage(img)
    # img[img>0.5] = 1
    # img[img<0.5] = 0
    img.astype(np.uint8)
    img = sitk.GetImageFromArray(img)
    img_save = sitk.BinaryDilate(img,(1,1,1)) # 膨胀
    img_save.SetSpacing(spacing)
    img_save.SetOrigin(origin)
    img_save.SetDirection(transfmat)
    sitk.WriteImage(img_save, savepath)

def dcm2nii_sitk(path_read, path_save):
    reader = sitk.ImageSeriesReader()
    seriesIDs = reader.GetGDCMSeriesIDs(path_read)
    N = len(seriesIDs)
    lens = np.zeros([N])
    for i in range(N):
        dicom_names = reader.GetGDCMSeriesFileNames(path_read, seriesIDs[i])
        lens[i] = len(dicom_names)
    N_MAX = np.argmax(lens)
    dicom_names = reader.GetGDCMSeriesFileNames(path_read, seriesIDs[N_MAX])
    reader.SetFileNames(dicom_names)
    image = reader.Execute()
    if not os.path.exists(path_save):
        os.mkdir(path_save)
    sitk.WriteImage(image, path_save+'/data.nii.gz')

def img_mask_consistent(imageFilepath,maskFilepath,new_maskFilepath):
    '''
      imageFilepath:
      outimageFilepath:
      new_spacing: [n,n,n]
      new_spacing: x,y,z
      is_label: if True, using Interpolator `sitk.sitkNearestNeighbor`
    '''
    image = sitk.ReadImage(imageFilepath) #读原图
    spacing = image.GetSpacing()
    origin = image.GetOrigin()
    transfmat = image.GetDirection()

    mask_itk = sitk.ReadImage(maskFilepath)
    mask_itk.SetSpacing(spacing)
    mask_itk.SetOrigin(origin)
    mask_itk.SetDirection(transfmat)
    sitk.WriteImage(mask_itk, new_maskFilepath)

# 提取np的非零值
def extract_npnozero_2d(data):
    outdata = []
    tmp = np.nonzero(data)# 返回非零索引
    for i in range(len(tmp[0])):
        outdata.append(data[tmp[0][i]][tmp[1][i]])
    return np.array(outdata)

# 获取img的spacing
def Getspacing(imageFilepath):
    '''
      imageFilepath:
      outimageFilepath:
      new_spacing: [n,n,n]
      new_spacing: x,y,z
      is_label: if True, using Interpolator `sitk.sitkNearestNeighbor`
    '''
    image = sitk.ReadImage(imageFilepath) #读原图
    spacing = np.array(image.GetSpacing())#读取原图spacing
    return spacing

# Resamping 重采样
def Changespacing(image,outimageFilepath,new_spacing=[1.0,1.0,1.0],is_label = False):
    '''
      image:
      outimageFilepath:
      new_spacing: [n,n,n]
      new_spacing: x,y,z
      is_label: if True, using Interpolator `sitk.sitkNearestNeighbor`
    '''
    size = np.array(image.GetSize())#读取原图尺寸
    spacing = np.array(image.GetSpacing())#读取原图spacing
    new_spacing = np.array(new_spacing)
    new_size = size * spacing  / new_spacing #计算新尺寸
    new_spacing_refine = size * spacing / new_size #计算新spacing
    new_spacing_refine = [float(s) for s in new_spacing_refine]
    new_size = [int(s) for s in new_size]

    resample = sitk.ResampleImageFilter()
    resample.SetOutputDirection(image.GetDirection())
    resample.SetOutputOrigin(image.GetOrigin())
    resample.SetSize(new_size)
    resample.SetOutputSpacing(new_spacing_refine)
    if is_label:
        resample.SetInterpolator(sitk.sitkNearestNeighbor)
    else:
        # resample.SetInterpolator(sitk.sitkBSpline)
        resample.SetInterpolator(sitk.sitkLinear)
    newimage = resample.Execute(image)
    sitk.WriteImage(newimage, outimageFilepath)

#外扩取小图
def Array_crop(nii_path,mask_path,waikuo=30):
    subject1 = wama()
    subject1.appendImageFromNifti('IMG',nii_path) #加载图像,自定义名
    subject1.appendSementicMaskFromNifti('IMG',mask_path) #加载mask
    img_box=subject1.getBbox('IMG')
    # 谁是x谁是z与读法有关,注意检查
    xx = img_box[1] - img_box[0]
    yy = img_box[3] - img_box[2]
    zz = img_box[5] - img_box[4]
    # 读取nii的其他信息
    mask_itk = sitk.ReadImage(nii_path)
    spacing = mask_itk.GetSpacing()
    origin = mask_itk.GetOrigin()
    transfmat = mask_itk.GetDirection()
    allpath =[nii_path,mask_path]
    flag = 0
    for path in allpath:
        flag += 1
        mask_itk = sitk.ReadImage(path)
        mask_img = sitk.GetArrayFromImage(mask_itk)
        # 调整窗宽窗位
        windows_center = 50
        windows_width = 150
        if flag == 1:
            min = windows_center - windows_width/2
            max = windows_center + windows_width/2
            mask_img[mask_img < min] = min
            mask_img[mask_img > max] = max
        small_img = np.zeros([zz,xx+waikuo,yy+waikuo])
        # small_img_array = mask_img[img_box[4]:img_box[5]+1,img_box[0]-int(waikuo/2):img_box[1]+int(waikuo/2)+1,img_box[2]-int(waikuo/2):img_box[3]+int(waikuo/2)+1]
        small_img_array = mask_img
        small_img = sitk.GetImageFromArray(small_img_array)
        small_img.SetSpacing(spacing)
        small_img.SetOrigin(origin)
        small_img.SetDirection(transfmat)

        if path == nii_path:
            small_mri = small_img
        else:
            small_roi = small_img
    return small_mri,small_roi

#将3D图像resize
def resize3D(img, aimsize, order=3):
        """
        :param img: 3D array
        :param aimsize: list, one or three elements, like [256], or [256,56,56]
        :return:
        """
        _shape = img.shape
        if len(aimsize) == 1:
            aimsize = [aimsize[0] for _ in range(3)]
        if aimsize[0] is None:
            return zoom(img, (1, aimsize[1] / _shape[1], aimsize[2] / _shape[2]), order=order)  # resample for cube_size
        if aimsize[1] is None:
            return zoom(img, (aimsize[0] / _shape[0], 1, aimsize[2] / _shape[2]), order=order)  # resample for cube_size
        if aimsize[2] is None:
            return zoom(img, (aimsize[0] / _shape[0], aimsize[1] / _shape[1], 1), order=order)  # resample for cube_size
        return zoom(img, (aimsize[0] / _shape[0], aimsize[1] / _shape[1], aimsize[2] / _shape[2]),
                    order=order)  # resample for cube_size

#比较小的图像补0
def add_zero(img, target_size=[6, 6]):
        zz, xx, yy = img.shape
        dim_min_list = [int(target_size[0] / 2 - xx / 2), int(target_size[1] / 2 - yy / 2)]
        target_size_img = np.zeros([zz, target_size[0], target_size[1]])
        target_size_img[:, dim_min_list[0]:dim_min_list[0] + xx, dim_min_list[1]:dim_min_list[1] + yy] = img
        return target_size_img


# 提取MRI特征
def cir_get_features(imageFilepath, maskFilepath, yamlpath):
    # 用yaml文件初始化特征提取器
    extractor = FEE.RadiomicsFeatureExtractor(yamlpath)  # todo 初始化特征提取器
    result = extractor.execute(imageFilepath, maskFilepath)
    feature = {}
    for key in result.keys():
        if not 'diagnostics' in key:
            feature[key] = float(result[key])
    return feature

# 提取MRI特征
def cir_get_featuresmap(imageFilepath, maskFilepath, yamlpath, name):
    # 用yaml文件初始化特征提取器
    extractor = FEE.RadiomicsFeatureExtractor(yamlpath)  # todo 初始化特征提取器
    result = extractor.execute(imageFilepath, maskFilepath)
    result1 = extractor.execute(imageFilepath, maskFilepath, voxelBased=True)
    feature = {}
    # 保存特征图
    feature1 = [[key, value] for key, value in result1.items() if not 'diagnostics' in key]
    listfeature = ["original_glszm_SizeZoneNonUniformity",
                   "wavelet-LLH_glcm_Imc2",
                   "wavelet-LLH_gldm_HighGrayLevelEmphasis",
                   "wavelet-LHL_glrlm_LowGrayLevelRunEmphasis",
                   "wavelet-LHH_gldm_DependenceNonUniformityNormalized",
                   "wavelet-LHH_gldm_LowGrayLevelEmphasis",
                   "wavelet-HLL_glrlm_ShortRunLowGrayLevelEmphasis",
                   "wavelet-HLH_glcm_MaximumProbability",
                   "wavelet-HLH_glrlm_LongRunLowGrayLevelEmphasis",
                   "wavelet-HHL_glszm_SizeZoneNonUniformity",
                   "wavelet-HHL_gldm_LargeDependenceHighGrayLevelEmphasis",
                   "wavelet-HHH_glcm_InverseVariance",
                   "wavelet-HHH_glrlm_GrayLevelNonUniformity",
                   "wavelet-HHH_glrlm_GrayLevelVariance",
                   "wavelet-HHH_glszm_SizeZoneNonUniformityNormalized"]

    for i in listfeature:
        flag = 0
        for x in feature1:
            if i == x[0]:
                image = feature1[flag][1]
                featurename = feature1[flag][0]
                #scan = sitk.GetArrayFromImage(image)
                print(featurename)
                sitk.WriteImage(image=image, fileName=r'Z:\data\FYH\LWY\experiment\EX_threecenter_new\Radiomic Feature Map\Map' + sep +
                                         name.split('.')[0] + '_' + featurename + '.nii.gz')
            flag += 1

    for key in result.keys():
        if not 'diagnostics' in key:
            feature[key] = float(result[key])
    return feature

def one_hot(array):
    from keras.utils.np_utils import to_categorical
    xx = to_categorical(array,num_classes=None)
    return xx

# if __name__=='__main__':
#     one_hot()